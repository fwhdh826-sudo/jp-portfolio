#!/usr/bin/env python3
"""
P5-B004b: JPX公式 東証上場銘柄一覧（listed-issues workbook）をprimary source
とするwhole-market universe providerと、eligibility v1 filter・cache/fallback・
provenanceを含む正規化されたprovider result契約。

P5-B005-JPX-UNIVERSE-PRODUCTION-RECOVERY: JPX公式CMSの添付ファイル形式が
data_j.xls（OLE2）→ data_j.xlsx（OOXML/ZIP）へ移行し、旧固定URLが404化した
ことへの回復。固定添付URLを別の固定URLへ差し替えるのではなく、公式の
上場銘柄一覧ページ（01.html）を起点に現行workbookリンクを都度discoverし、
URL authority（scheme/host/namespace/basename/extension）をstrictに検証した
うえで、bytesシグネチャからxls/xlsxを判定してparseする。legacy xls経路
（xlrd）は温存し、xlsx経路（openpyxl, read_only/data_only）を同じ正規化row
契約へ合流させる。

責務:
  - JPX公式listing pageをHTTPS取得し、現行listed-issues workbookリンクを
    discover・authority検証（fetch_listing_page/discover_workbook_url）
  - discoverしたworkbookをHTTPS取得し、bytesシグネチャでxls(OLE2)/xlsx(ZIP)
    を判定してparse（detect_workbook_format/parse_jpx_workbook_bytes）。
    extension↔signature不一致・HTMLエラーページはfail closed
  - eligibility v1 filter（プライム内国株式のみ、5桁優先株/種類株式除外）
  - internal-onlyなlast-good cacheと、fetch/parse/schema/duplicate/
    row-count異常時のcache→seed_list_v1の3段fallback chain
  - provenance情報を含む正規化されたprovider result契約(JPXUniverseResult)

非範囲（B004cへ持ち越し）:
  - cheap pre-screen（eligible universe（約1552件）をMAX_ENRICHMENT_UNIVERSE=
    500以下のbounded shortlistへ絞る処理）は本モジュールでは実装しない。
  - JPXUniverseResultをdata/build_candidates_stocks.pyの
    default_universe_provider()として本番接続すること。このticketでは
    provider/cache/fallbackの実装までであり、既存のSEED_LIST(41銘柄)による
    本番挙動には一切影響しない。

privacy: 本モジュールが扱うのはJPXが公開する市場情報（銘柄コード・銘柄名・
市場区分・業種区分）のみであり、保有・取引・口座・現金等の個人情報は一切
参照・保存しない。cacheは internal only（data/.jpx_cache/ 配下）であり、
public/data配下へは一切コピーしない。

1559 vs 1560件差の説明（P5-B004b-1 dry-run vs Fable監査、2026-07-14実データで検証済み）:
  dry-runは「市場・商品区分」列に対し文字列包含判定
  `df[col].str.contains("プライム")` を用いていたため、
  「プライム（内国株式）」1559件に加え「プライム（外国株式）」1件
  （2026-07-14時点データでは同区分は1銘柄のみ存在）を誤って含み、
  合計1560件になっていた。Fable監査の1559件は「プライム（内国株式）」の
  厳密一致による件数であり、本providerもeligibility v1として
  market_segment の厳密一致（プライム（内国株式）のみ）を採用するため、
  1559件が正しいPrime内国株式候補数となる。
  さらにeligibility v1では5桁コード（優先株/種類株式、7件）を除外するため、
  最終eligibleCountは1552件になる（詳細はapply_eligibility()のdocstring参照）。
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from datetime import date, datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, NamedTuple
from urllib.parse import urljoin, urlsplit

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------

# 公式の上場銘柄一覧authority page（安定URL）。ここから現行workbookリンクを
# discoverする。opaqueなCMS添付ID（tvdivq0000001vg2-att）が恒久である前提は
# 置かない。
JPX_LISTING_PAGE_URL = (
    "https://www.jpx.co.jp/markets/statistics-equities/misc/01.html"
)
# listing page URL（初期URL・redirect先いずれも）が満たすべき正規path。
# www.jpx.co.jp配下の任意のpath（/corporate/・/english/・
# /markets/statistics-equities/ 配下の別page等）を許可するのは広すぎる
# authorityであり、CMS内の無関係pageへのredirectをfetchしてしまう
# （P5-B005-R3 Blocker A）。narrowest correct contractとして、この
# 上場銘柄一覧authority page 1つのnormalized pathのみを許可する。
JPX_LISTING_PAGE_PATH = "/markets/statistics-equities/misc/01.html"

# discoverしたworkbook URLが満たすべきauthority制約。
JPX_ALLOWED_SCHEME = "https"
JPX_ALLOWED_HOST = "www.jpx.co.jp"
# 添付ファイルは必ずこのnamespace配下に置かれる。
JPX_ATTACHMENT_PATH_PREFIX = "/markets/statistics-equities/misc/"
# CMS添付コンテナディレクトリは "-att" で終わる（例: tvdivq0000001vg2-att）。
JPX_ATTACHMENT_DIR_SUFFIX = "-att"
# 許可するlisted-issues workbook basename（xls/xlsx両対応）。
APPROVED_WORKBOOK_BASENAMES = ("data_j.xls", "data_j.xlsx")
APPROVED_WORKBOOK_EXTENSIONS = (".xls", ".xlsx")

# legacy固定添付URL（現在404）。default fetch pathでは使わない。
# fetch_jpx_xls() のdefault引数としてのみ残す（後方互換）。
JPX_SOURCE_URL = (
    "https://www.jpx.co.jp/markets/statistics-equities/misc/"
    "tvdivq0000001vg2-att/data_j.xls"
)
# 公開provenance上のsource識別子。workbook形式（xls/xlsx）に依存しない安定値
# （形式差はJPXUniverseResult.workbook_format / 内部ログで区別する。§14）。
SOURCE_IDENTIFIER = "jpx_data_j_xls"
FETCH_TIMEOUT_SECONDS = 30

UNIVERSE_ID = "jpx_prime_domestic_v1"
FALLBACK_UNIVERSE_ID = "seed_list_v1"

REQUIRED_COLUMNS = ("コード", "銘柄名", "市場・商品区分")
# xls: OLE2 / Compound File Binary。xlsx: ZIP local file header（PK\x03\x04）。
OLE2_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
ZIP_LOCAL_FILE_SIGNATURE = b"PK\x03\x04"
ZIP_EMPTY_SIGNATURE = b"PK\x05\x06"
ZIP_SPANNED_SIGNATURE = b"PK\x07\x08"

# eligibility v1: プライムかつ内国株式（普通株式相当）のみを厳密一致で採用する。
# ETF・ETN / REIT・ベンチャーファンド等 / PRO Market / 出資証券 / 外国株式 /
# Standard・Growthは、いずれもこの値と異なるためこの1回のマッチで自動的に
# 除外される（2026-07-14実データで確認した「市場・商品区分」列の値は
# スタンダード（内国株式）/プライム（内国株式）/グロース（内国株式）/
# ETF・ETN/PRO Market/REIT・ベンチャーファンド・カントリーファンド・
# インフラファンド/スタンダード（外国株式）/グロース（外国株式）/出資証券/
# プライム（外国株式）の10種）。
MARKET_SEGMENT_PRIME_DOMESTIC = "プライム（内国株式）"

# 2026-07-14実データで確認したJPX「市場・商品区分」列の全既知値
# （上記10種、モジュールdocstring参照）。segment_countsのkeyがこの
# 集合の外に出ることは、JPXが市場区分を追加/変更したか、cache
# payloadが改竄されたかのいずれかであり、cache authorityとしては
# 承認済み集合の範囲内でのみ受理する（P5-B005-R3 §15）。
APPROVED_MARKET_SEGMENTS = frozenset({
    "スタンダード（内国株式）",
    MARKET_SEGMENT_PRIME_DOMESTIC,
    "グロース（内国株式）",
    "ETF・ETN",
    "PRO Market",
    "REIT・ベンチャーファンド・カントリーファンド・インフラファンド",
    "スタンダード（外国株式）",
    "グロース（外国株式）",
    "出資証券",
    "プライム（外国株式）",
})

# apply_eligibility()が返すfilters_appliedの各stage名（writer authority）。
# cache validator（_cache_authority_valid）はこれらの定数を直接参照して
# 検証することで、writerとvalidatorの間でstage名の定義が二重化・
# 乖離することを防ぐ（P5-B005-R3 §11/§16）。
FILTER_STAGE_SOURCE_ROWS = "source_rows"
FILTER_STAGE_MARKET_SEGMENT_PRIME_DOMESTIC = (
    "market_segment_prime_domestic_common_strict_match"
)
FILTER_STAGE_EXCLUDE_PREFERRED_OR_CLASS = (
    "exclude_preferred_or_class_shares_5digit_code"
)
FILTERS_APPLIED_STAGE_ORDER = (
    FILTER_STAGE_SOURCE_ROWS,
    FILTER_STAGE_MARKET_SEGMENT_PRIME_DOMESTIC,
    FILTER_STAGE_EXCLUDE_PREFERRED_OR_CLASS,
)

# 前回正常row_countの何%未満ならsource異常とみなしcache fallbackへ回すか。
ROW_COUNT_MIN_RATIO = 0.7

# 絶対floor（前回cacheの有無に関わらず常に適用）。2026-07-14実データでは
# raw row_count=4437・eligible_count=1552。将来の銘柄数増減を過剰に拒否
# しないよう実データの1/4程度に大きな余裕を持たせつつ、HTMLエラーページの
# 誤parseやfirst-run（前回cacheなし）でのtruncated sourceをbaselineとして
# 採用してしまう事態（F7）を防ぐ安全弁として機能する。
MIN_RAW_ROW_COUNT = 1000
MIN_ELIGIBLE_COUNT = 300

# 前回last-good cacheのeligible件数（=cacheのitems件数）の何%未満なら
# eligibility異常（market label drift等）とみなしcache fallbackへ回すか。
# raw row_countは正常でもeligible_countだけが急減するケース（F1）を検知する
# ため、raw用のROW_COUNT_MIN_RATIOとは独立に判定する。
ELIGIBLE_COUNT_MIN_RATIO = 0.7

CACHE_PATH = Path(__file__).parent / ".jpx_cache" / "jpx_universe_cache.json"
CACHE_SCHEMA_KIND = "jpx_universe_cache_v1"

# P5-B005-R4: current-run cache attestation（ephemeral、§17-21）。
# jpx_universe_cache.json（persistent last-good cache、actions/cacheで
# cross-run永続化される）とは別の、"このrunner・このrunが実際にこの
# cache bytesをlive JPXから生成した"ことだけを証明するsidecar。
# data/.jpx_cache/ 配下（.gitignore対象、CACHE_PATH参照）に置くことで
# 追加のgitignoreルールなしでcommit対象外になる。actions/cacheの
# path（full_batch.yml）はjpx_universe_cache.jsonの1ファイルのみを
# 指すため、このsidecarはcross-run永続化されない
# （毎runで揮発する——これは意図的な設計であり欠陥ではない、§21）。
ATTESTATION_PATH = (
    Path(__file__).parent / ".jpx_cache" / "jpx_universe_cache.attestation.json"
)
ATTESTATION_SCHEMA_KIND = "jpx_universe_cache_attestation_v1"

# last-good cacheとして使ってよい最大経過時間。Full Batchは平日毎朝走るが、
# 複数日にわたるJPX live取得断絶でもfallbackとして使えるよう十分な余裕を
# 持たせつつ、10年前cache等の明らかにstaleなpayloadは拒否できる値として
# 30日（720h）を採用する（P5-B005-R2: RESTORED_CACHE_REVALIDATED §16）。
MAX_CACHE_AGE_HOURS = 24 * 30
# fetched_atがnowよりわずかに未来（save直後の時計ずれ等）でも許容する
# clock-skew tolerance。これを超える未来日時のcacheは拒否する。
CACHE_CLOCK_SKEW_TOLERANCE_HOURS = 1.0

# workbook URL探索でfollowするredirectの最大段数。各hopのLocationは
# 次のrequestを出す前に必ずJPX authorityへ再検証する（P5-B005-R2 §9）。
MAX_REDIRECTS = 5


# ---------------------------------------------------------------------------
# 例外
# ---------------------------------------------------------------------------


class JPXFetchError(RuntimeError):
    """HTTPS取得失敗、非200応答、想定外シグネチャ（OLE2/xls以外）。"""


class JPXParseError(RuntimeError):
    """xlrdによるworkbook/parse失敗。"""


class JPXSchemaError(RuntimeError):
    """必須列欠損、重複code検出等のschema異常。"""


class JPXRowCountGuardError(RuntimeError):
    """raw row_countが異常（絶対floor未満、または前回正常値の
    ROW_COUNT_MIN_RATIO未満に急減）。"""


class JPXEligibleCountGuardError(RuntimeError):
    """eligible_countが異常（絶対floor未満、または前回last-good cacheの
    eligible件数のELIGIBLE_COUNT_MIN_RATIO未満に急減）。raw row_countは
    正常でもmarket label drift等でeligible側だけが崩壊するケース（F1）を
    捕捉する。"""


# ---------------------------------------------------------------------------
# 正規化されたprovider result契約
# ---------------------------------------------------------------------------


class JPXUniverseResult(NamedTuple):
    """JPX universe providerの戻り値。既存data/build_candidates_stocks.pyの
    UniverseResult(universe_id, items)契約を踏襲しつつ、provenance/
    eligibility/fallback情報を追加した正規化contract。

    itemsは既存UniverseResult.itemsと同じ (code, name, sector) タプルの
    リスト形式（sectorはJPX「33業種区分」列を採用）。

    注意: このticketではJPXUniverseResultを
    data/build_candidates_stocks.pyのUniverseProvider
    （Callable[[], UniverseResult]）として直接接続しない
    （B004cでpre-screen後に接続する）。UniverseResult自体は変更しない
    ——既存41 seed production behavior・既存testsに一切影響しない。
    """

    universe_id: str
    items: list[tuple[str, str, str]]
    source: str
    source_identifier: str
    fetched_at: str
    source_as_of: str | None
    row_count: int
    eligible_count: int
    segment_counts: dict[str, int]
    filters_applied: list[dict[str, Any]]
    fallback_used: bool
    cache_age_hours: float | None
    dropped_rows: list[dict[str, Any]]
    # "xls" / "xlsx" / None（cache・seed fallback時、または形式判定不能時）。
    # 公開candidates_stocks schemaには出さない内部provenance（§14）。末尾の
    # optionalフィールドとして追加——既存のkeyword構築・duck typing consumerに
    # 影響しない。
    workbook_format: str | None = None


# ---------------------------------------------------------------------------
# Source adapter: official listing-page discovery（§6/§7/§8）
# ---------------------------------------------------------------------------


class _AnchorHrefExtractor(HTMLParser):
    """<a ... href="..."> のanchor要素のhref属性値のみを抽出するHTMLParser。

    html.parser.HTMLParserの構文解析に委ねることで、以下は構造的に
    href候補へ混入しない（正規表現による生テキストscanの弱点を排除する。
    P5-B005-R2: DISCOVERY_HTML_ROBUSTNESS）:
      - data-href等、"href"を含むが別名の属性（handle_starttagはtag名と
        属性名の完全一致でのみ拾う）
      - HTMLコメント内の文字列（コメントはhandle_commentへ渡り、
        handle_starttagは呼ばれない）
      - <script>内のテキスト（script要素の中身はCDATA相当として扱われ、
        タグとしてparseされない）
      - 地の文（plain text）中の"href=..."という文字列そのもの

    HTMLParserは既定でattribute値の文字参照（entity）をunescapeして
    渡すため、ここで再度html.unescape()を呼ぶと二重decodeになる
    （実機動作で確認済み。本ticketではattrsの値をそのまま使う）。"""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.hrefs: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        for name, value in attrs:
            if name.lower() == "href" and value is not None:
                self.hrefs.append(value)


def _extract_anchor_hrefs(page_html: str) -> list[str]:
    """page_html中の実際の<a href="...">属性値のみを順序維持で返す。
    malformed HTML（未閉タグ等）でもHTMLParserは可能な範囲でbest-effort
    parseを続ける——それ自体は検出漏れを増やさない（承認済みbasenameへの
    一致・authority検証は後段で別途行うため、ここでは抽出のみを担う）。"""
    parser = _AnchorHrefExtractor()
    parser.feed(page_html)
    parser.close()
    return parser.hrefs


def _bounded_redirect_fetch(
    initial_url: str,
    validate_fn: Any,
    timeout: int,
    max_redirects: int = MAX_REDIRECTS,
) -> tuple[Any, str]:
    """allow_redirects=Falseで手動bounded-redirectを行うHTTPS GET helper。

    どのURL（初期URL・各redirect Location）も、requestsで実際にrequestを
    出す前に必ずvalidate_fn()で検証する——redirect先のbodyは、そのURLが
    検証を通過するまで一切consumeしない（P5-B005-R2:
    REDIRECT_AUTHORITY_REVALIDATED）。

    拒否: Locationヘッダ欠如・redirect loop（同一URL再訪）・
    max_redirects超過・validate_fn()が拒否するforeign/downgrade/
    traversal先。最終的な200応答のresp.urlも再度validate_fn()へ通してから
    返す（redirect無しの直接200応答でも同じ経路で再検証する）。"""
    try:
        import requests
    except ImportError as e:
        raise JPXFetchError(f"requests is not installed: {e!r}") from e

    seen: set[str] = set()
    url = validate_fn(initial_url)
    for _ in range(max_redirects + 1):
        if url in seen:
            raise JPXFetchError(f"redirect loop detected at {url!r}")
        seen.add(url)

        try:
            resp = requests.get(url, timeout=timeout, allow_redirects=False)
        except Exception as e:  # noqa: BLE001
            raise JPXFetchError(f"fetch failed: {e!r}") from e

        if resp.status_code in (301, 302, 303, 307, 308):
            location = resp.headers.get("Location")
            if not location:
                raise JPXFetchError(f"redirect without Location header from {url!r}")
            next_url = urljoin(url, location)
            url = validate_fn(next_url)
            continue

        if resp.status_code != 200:
            raise JPXFetchError(f"unexpected status_code={resp.status_code} at {url!r}")

        final_url = validate_fn(str(resp.url))
        return resp, final_url

    raise JPXFetchError(f"too many redirects (>{max_redirects}) starting at {initial_url!r}")


def fetch_listing_page(
    url: str = JPX_LISTING_PAGE_URL, timeout: int = FETCH_TIMEOUT_SECONDS
) -> str:
    """公式の上場銘柄一覧authority page（01.html）を取得しHTML textを返す。
    非200・非HTML・network例外・requests未導入はいずれもJPXFetchErrorへ
    正規化する（呼び出し側でcache/seed fallbackへ回す）。

    redirectはallow_redirects=Falseで手動bounded-followし、各hopのLocationを
    _validate_listing_page_url()で再検証してから次のrequestを出す
    （P5-B005-R2: DISCOVERED_URL_SECURITY §9/§10）。"""
    resp, _final_url = _bounded_redirect_fetch(url, _validate_listing_page_url, timeout)

    ctype = resp.headers.get("content-type", "")
    if "html" not in ctype.lower():
        raise JPXFetchError(f"listing page content-type is not HTML: {ctype!r}")

    return resp.text


def _reject_encoded_path_escape(path: str, what: str) -> None:
    """pathに"%"が1文字でも含まれる場合はfail closedで拒否する。

    JPX workbook attachment path（/markets/statistics-equities/misc/
    <...-att>/data_j.(xls|xlsx)）とlisting page pathはいずれもASCIIの
    固定文字集合のみで構成され、percent-encodingを正当に必要とする
    セグメントは存在しない。%2e%2e%2f・%2E%2E%5C・%252e%252e%252f
    （二重encode）はいずれも"%"を含むため、decode/canonicalizeを一切
    行わずこの1判定だけでtraversal escapeをfail closedに拒否できる
    （単純unquoteによる二重encode見逃しを避ける。P5-B005-R2 §12）。"""
    if "%" in path:
        raise JPXFetchError(f"percent-encoded characters not allowed in {what}: {path!r}")


def _validate_listing_page_url(url: str) -> str:
    """listing page URL（初期URL・redirect先いずれも）がJPX listing-page
    authority契約を満たすか検証する（P5-B005-R3 Blocker A）。

    許可: https / host==www.jpx.co.jp（userinfoなし） / query・fragment
    なし / pathが承認済みJPX統計・上場銘柄一覧authority page
    （JPX_LISTING_PAGE_PATH）に正規化後exact一致。
    拒否: http・別origin・userinfo trick・query/fragment付与・
    path traversal（literal/encoded問わず）・www.jpx.co.jp配下の
    他の任意path（/corporate/・/english/・
    /markets/statistics-equities/ 配下の別pageを含む）。
    "同一host"だけではauthorityとして不十分——workbook URLと同じ
    strictさでnamespaceをexact一致まで絞る（§7-8）。"""
    if not isinstance(url, str) or not url.strip():
        raise JPXFetchError("empty listing page URL")

    parts = urlsplit(url)
    if parts.scheme != JPX_ALLOWED_SCHEME:
        raise JPXFetchError(f"non-https listing page URL scheme: {parts.scheme!r}")
    # netloc（userinfo付き含む）の完全一致要求により、
    # "user@www.jpx.co.jp@evil.example" 等のuserinfo trickも
    # www.jpx.co.jp単体のnetlocとは一致せず自動的に拒否される。
    if parts.netloc != JPX_ALLOWED_HOST:
        raise JPXFetchError(f"unexpected listing page host: {parts.netloc!r}")
    if parts.query or parts.fragment:
        raise JPXFetchError("listing page URL must not carry query/fragment")

    _reject_encoded_path_escape(parts.path, "listing page path")
    segments = parts.path.split("/")
    if any(seg in (".", "..") for seg in segments):
        raise JPXFetchError(f"path traversal in listing page path: {parts.path!r}")

    if parts.path != JPX_LISTING_PAGE_PATH:
        raise JPXFetchError(
            f"listing page path outside approved authority "
            f"{JPX_LISTING_PAGE_PATH!r}: {parts.path!r}"
        )

    return url


def _validate_workbook_authority(resolved: str) -> str:
    """解決済み絶対URLに対してworkbook authority制約（§7）をstrictに検証
    する。呼び出し元は以下の2箇所:
      - _validate_discovered_workbook_url(): listing page hrefをurljoinで
        絶対化した直後
      - download_workbook()の_bounded_redirect_fetch(): 最初のURLと、
        redirectで示された各Location（urljoin後）を再検証

    許可: https / host==www.jpx.co.jp / path が
    /markets/statistics-equities/misc/<...-att>/data_j.(xls|xlsx) の形
    （query/fragmentなし、encoded文字なし、"." ".." セグメントなし、
    "//" なし）。
    拒否: http・別origin・userinfo trick・path traversal（literal/encoded/
    二重encoded問わず）・想定外basename・想定外namespace・
    query/fragment authority trick。"""
    parts = urlsplit(resolved)

    if parts.scheme != JPX_ALLOWED_SCHEME:
        raise JPXFetchError(f"non-https workbook URL scheme: {parts.scheme!r}")
    if parts.netloc != JPX_ALLOWED_HOST:
        raise JPXFetchError(f"unexpected workbook host: {parts.netloc!r}")
    if parts.query or parts.fragment:
        raise JPXFetchError("workbook URL must not carry query/fragment")

    path = parts.path
    _reject_encoded_path_escape(path, "workbook path")
    if "//" in path:
        raise JPXFetchError(f"malformed workbook path: {path!r}")
    segments = path.split("/")
    if any(seg in (".", "..") for seg in segments):
        raise JPXFetchError(f"path traversal in resolved workbook path: {path!r}")
    if "" in segments[1:]:
        raise JPXFetchError(f"empty segment in workbook path: {path!r}")

    if not path.startswith(JPX_ATTACHMENT_PATH_PREFIX):
        raise JPXFetchError(
            f"workbook path outside allowed namespace {JPX_ATTACHMENT_PATH_PREFIX!r}: {path!r}"
        )
    if len(segments) < 6:
        raise JPXFetchError(f"workbook path too shallow for an attachment: {path!r}")

    basename = segments[-1]
    container = segments[-2]
    if basename not in APPROVED_WORKBOOK_BASENAMES:
        raise JPXFetchError(f"unexpected workbook basename: {basename!r}")
    if not basename.endswith(APPROVED_WORKBOOK_EXTENSIONS):
        raise JPXFetchError(f"unexpected workbook extension: {basename!r}")
    if not container.endswith(JPX_ATTACHMENT_DIR_SUFFIX):
        raise JPXFetchError(
            f"workbook not under a CMS attachment container ({JPX_ATTACHMENT_DIR_SUFFIX!r}): {path!r}"
        )

    return resolved


def _validate_discovered_workbook_url(
    href: str, base_url: str = JPX_LISTING_PAGE_URL
) -> str:
    """listing pageから抽出したhrefを絶対URLへ解決し、
    _validate_workbook_authority()へ委譲する。満たさない場合はJPXFetchError。"""
    if not isinstance(href, str) or not href.strip():
        raise JPXFetchError("empty workbook href")
    raw = href.strip()

    lowered = raw.lower()
    for bad_scheme in ("javascript:", "data:", "vbscript:", "file:", "about:"):
        if lowered.startswith(bad_scheme):
            raise JPXFetchError(f"rejected workbook href scheme: {raw!r}")

    # urljoin前の生hrefでの明示的なtraversal拒否（解決後のnamespace checkでも
    # 弾かれるが、理由を明確にするため前段で落とす）。
    if ".." in re.split(r"[\\/]", raw):
        raise JPXFetchError(f"path traversal in workbook href: {raw!r}")
    if "\\" in raw:
        raise JPXFetchError(f"backslash in workbook href: {raw!r}")

    resolved = urljoin(base_url, raw)
    return _validate_workbook_authority(resolved)


def discover_workbook_url(
    page_html: str, base_url: str = JPX_LISTING_PAGE_URL
) -> str:
    """listing page HTMLから現行listed-issues workbookリンクをdiscoverする。

    候補hrefは実際の<a href="...">属性値のみ（_extract_anchor_hrefs、
    html.parser.HTMLParserベース）とし、approved basename
    （data_j.xls / data_j.xlsx）を指すものだけをそれぞれ
    _validate_discovered_workbook_url()へ通す。検証を通る一意なURLが
    得られた場合のみそれを返す。0件・parse不能・複数の相異なる
    approved URL（xls/xlsx混在など曖昧な状態）はいずれもfail closed
    （JPXFetchError）——推測した固定URLへdefaultしない（§8）。"""
    if not isinstance(page_html, str) or not page_html.strip():
        raise JPXFetchError("empty listing page HTML")

    hrefs = _extract_anchor_hrefs(page_html)

    accepted: list[str] = []
    for href in hrefs:
        tail = href.split("#", 1)[0].split("?", 1)[0].strip()
        basename = tail.rsplit("/", 1)[-1].lower()
        if basename not in APPROVED_WORKBOOK_BASENAMES:
            continue
        try:
            accepted.append(_validate_discovered_workbook_url(href, base_url))
        except JPXFetchError:
            # approved basenameだがauthority検証に落ちたhrefは黙って捨てる
            # （untrusted linkをpickしない）。
            continue

    unique = sorted(set(accepted))
    if not unique:
        raise JPXFetchError(
            "no approved listed-issues workbook link found on JPX listing page"
        )
    if len(unique) > 1:
        raise JPXFetchError(
            f"ambiguous approved workbook links on JPX listing page: {unique}"
        )
    return unique[0]


def detect_workbook_format(content: bytes) -> str:
    """downloadしたbytesのシグネチャからworkbook形式を判定する。
    "xls"（OLE2/Compound File）または "xlsx"（ZIP/OOXML）。

    HTMLエラーページ・truncated/garbage・空/spanned ZIPはJPXFetchError。
    拡張子ではなく必ずbytesで判定する（§9）。"""
    if not isinstance(content, (bytes, bytearray)):
        raise JPXFetchError(f"workbook content is not bytes: {type(content).__name__}")
    if len(content) < 8:
        raise JPXFetchError(f"workbook content too small: {len(content)} bytes")

    if content.startswith(OLE2_SIGNATURE):
        return "xls"
    if content.startswith(ZIP_LOCAL_FILE_SIGNATURE):
        return "xlsx"
    if content.startswith((ZIP_EMPTY_SIGNATURE, ZIP_SPANNED_SIGNATURE)):
        raise JPXFetchError("ZIP archive is empty or spanned, not an xlsx workbook")

    head = content[:512].lstrip().lower()
    if head.startswith((b"<!doctype", b"<html", b"<?xml", b"<head", b"<body")):
        raise JPXFetchError("received HTML/XML error page, not a workbook")
    raise JPXFetchError(
        f"unrecognized workbook signature: {bytes(content[:8]).hex()!r}"
    )


def download_workbook(
    url: str, timeout: int = FETCH_TIMEOUT_SECONDS
) -> tuple[bytes, str]:
    """検証済みworkbook URLをHTTPS取得し、(content, format) を返す。

    非200・network例外・requests未導入はJPXFetchError。redirectは
    allow_redirects=Falseで手動bounded-followし、各hopのLocationおよび
    最終応答URLを_validate_workbook_authority()で再検証してから次の
    requestを出す／bodyをconsumeする（P5-B005-R2: DISCOVERED_URL_SECURITY
    §9/§11）。URL拡張子とbytesシグネチャが不一致（.xlsxなのにOLE2、.xls
    なのにZIP等）の場合はfail closed（§9）。extension判定は最終検証済みURL
    （リダイレクト後）に対して行う。"""
    resp, final_url = _bounded_redirect_fetch(url, _validate_workbook_authority, timeout)

    content = resp.content
    detected = detect_workbook_format(content)

    ext_format = "xlsx" if final_url.lower().endswith(".xlsx") else "xls"
    if ext_format != detected:
        raise JPXFetchError(
            f"workbook extension/signature mismatch: URL says .{ext_format} "
            f"but bytes are {detected}"
        )
    return content, detected


def fetch_jpx_workbook_bytes(timeout: int = FETCH_TIMEOUT_SECONDS) -> bytes:
    """default fetch path: listing page取得 → workbook URL discover →
    workbook download（signature検証込み）。bytesのみ返す
    （形式はparse_jpx_workbook_bytes / detect_workbook_format側で再判定する）。

    いずれの段階の失敗もJPXFetchErrorとして送出され、get_jpx_universe()の
    live → valid cache → seed_list_v1 fallback chainへ回る。"""
    html = fetch_listing_page(timeout=timeout)
    url = discover_workbook_url(html)
    content, _fmt = download_workbook(url, timeout=timeout)
    return content


def fetch_jpx_xls(url: str = JPX_SOURCE_URL, timeout: int = FETCH_TIMEOUT_SECONDS) -> bytes:
    """（legacy・後方互換）固定data_j.xls URLをHTTPS取得する。現在この添付URLは
    404であり、default fetch pathでは使わない（fetch_jpx_workbook_bytes参照）。

    200以外・OLE2/xls以外のシグネチャはJPXFetchErrorとして送出する。
    requests未導入（ImportError/ModuleNotFoundError）もfetch異常の一種として
    JPXFetchErrorへ正規化する。"""
    try:
        import requests
    except ImportError as e:
        raise JPXFetchError(f"requests is not installed: {e!r}") from e

    try:
        resp = requests.get(url, timeout=timeout)
    except Exception as e:  # noqa: BLE001 - network例外は全てfetch失敗として扱う
        raise JPXFetchError(f"fetch failed: {e!r}") from e

    if resp.status_code != 200:
        raise JPXFetchError(f"unexpected status_code={resp.status_code}")

    content = resp.content
    if not content.startswith(OLE2_SIGNATURE):
        raise JPXFetchError(
            f"unexpected file signature (not OLE2/xls): {content[:8].hex()!r}"
        )
    return content


# ---------------------------------------------------------------------------
# Source adapter: parse
# ---------------------------------------------------------------------------


def _code_to_str(raw: Any) -> str:
    """codeは必ずstrとして保持する。xlrdの数値セルはfloat、openpyxlの数値セルは
    intで返るため、整数値であればゼロ小数を落としてstr化し、文字列セル
    （英字混在code等）はstrip済みのstrをそのまま使う。boolは数値扱いしない。"""
    if isinstance(raw, bool):
        return str(raw)
    if isinstance(raw, int):
        return str(raw)
    if isinstance(raw, float):
        if raw.is_integer():
            return str(int(raw))
        return str(raw)
    return str(raw).strip()


def parse_rows_from_sheet(sheet: Any) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """xlrd sheet互換オブジェクト（.nrows/.ncols/.cell(r,c)で
    ctype/valueを持つセルを返す）からrowを抽出する。

    戻り値: (parsed_rows, dropped_rows)。
    parsed_rowsの各要素: {"code": str, "name": str, "market_segment": str,
    "sector": str, "source_as_of_raw": Any}
    dropped_rowsの各要素: {"row": int, "reason": str, "code": str | None}
    """
    if sheet.nrows == 0:
        raise JPXSchemaError("empty sheet (no header row)")

    header = [str(sheet.cell(0, c).value).strip() for c in range(sheet.ncols)]
    missing_cols = [c for c in REQUIRED_COLUMNS if c not in header]
    if missing_cols:
        raise JPXSchemaError(f"missing required columns: {missing_cols}")

    col_idx = {name: header.index(name) for name in header}
    code_c = col_idx["コード"]
    name_c = col_idx["銘柄名"]
    market_c = col_idx["市場・商品区分"]
    sector_c = col_idx.get("33業種区分")
    date_c = col_idx.get("日付")

    parsed: list[dict[str, Any]] = []
    dropped: list[dict[str, Any]] = []

    for r in range(1, sheet.nrows):
        code_raw = sheet.cell(r, code_c).value
        name_raw = sheet.cell(r, name_c).value
        market_raw = sheet.cell(r, market_c).value

        if code_raw is None or (isinstance(code_raw, str) and not code_raw.strip()):
            dropped.append({"row": r, "reason": "missing_code", "code": None})
            continue

        code = _code_to_str(code_raw)
        name = str(name_raw).strip() if name_raw is not None else ""
        market = str(market_raw).strip() if market_raw is not None else ""

        if not code or not name or not market:
            dropped.append({"row": r, "reason": "missing_required_field", "code": code or None})
            continue

        sector = ""
        if sector_c is not None:
            sector_raw = sheet.cell(r, sector_c).value
            sector = str(sector_raw).strip() if sector_raw is not None else ""

        source_as_of_raw = sheet.cell(r, date_c).value if date_c is not None else None

        parsed.append({
            "code": code,
            "name": name,
            "market_segment": market,
            "sector": sector,
            "source_as_of_raw": source_as_of_raw,
        })

    return parsed, dropped


def parse_jpx_xls_bytes(content: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """OLE2/xls bytesをxlrdでparseし、parse_rows_from_sheet()へ委譲する。

    xlrd自体が未導入の環境（ImportError/ModuleNotFoundError）、workbook open
    失敗、sheet_by_index等のparse adapter失敗は、いずれもJPXParseErrorへ
    正規化する。これによりdependency欠如を含むparse異常が
    live → valid cache → seed_list_v1 のfallback chainを突破しない
    （parse_rows_from_sheet自身が送出するJPXSchemaError/JPXSchemaError系の
    schema異常はそのまま呼び出し元へ伝播させる——parseは成功したがschemaが
    不正、という区別を保つため）。"""
    try:
        import xlrd
    except ImportError as e:
        raise JPXParseError(f"xlrd is not installed: {e!r}") from e

    try:
        wb = xlrd.open_workbook(file_contents=content)
        sheet = wb.sheet_by_index(0)
    except Exception as e:  # noqa: BLE001
        raise JPXParseError(f"xlrd parse failed: {e!r}") from e

    return parse_rows_from_sheet(sheet)


class _MatrixSheet:
    """openpyxl等が返す行列（list[tuple]）を、parse_rows_from_sheet()が期待する
    xlrd sheet互換interface（.nrows / .ncols / .cell(r, c).value）へ薄く適合させる。

    xls経路とxlsx経路を単一の正規化ロジック（parse_rows_from_sheet）へ合流させ、
    同一workbookのxls/xlsx入力が等価な正規化結果を返すことを保証する（§11/§25）。
    ragged row（列数不揃い）は範囲外セルをNone扱いにする。"""

    __slots__ = ("_matrix", "nrows", "ncols")

    def __init__(self, matrix: list) -> None:
        self._matrix = matrix
        self.nrows = len(matrix)
        self.ncols = len(matrix[0]) if matrix else 0

    def cell(self, r: int, c: int) -> Any:
        row = self._matrix[r]
        value = row[c] if 0 <= c < len(row) else None
        return _MatrixCell(value)


class _MatrixCell:
    __slots__ = ("value",)

    def __init__(self, value: Any) -> None:
        self.value = value


def parse_jpx_xlsx_bytes(content: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """ZIP/OOXML/xlsx bytesをopenpyxlでparseし、parse_rows_from_sheet()へ委譲する。

    openpyxlはread_only=True / data_only=Trueで開く（マクロ実行なし、数式は
    キャッシュ値のみ、大きなシートでも低メモリ）。openpyxl未導入・workbook open
    失敗・sheet読み出し失敗はいずれもJPXParseErrorへ正規化する
    （parse_rows_from_sheetが送出するschema系例外はそのまま伝播）。"""
    try:
        import openpyxl
    except ImportError as e:
        raise JPXParseError(f"openpyxl is not installed: {e!r}") from e

    import io

    wb = None
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
        if not wb.sheetnames:
            raise JPXParseError("xlsx workbook has no sheets")
        ws = wb[wb.sheetnames[0]]
        matrix = [tuple(row) for row in ws.iter_rows(values_only=True)]
    except JPXParseError:
        raise
    except Exception as e:  # noqa: BLE001
        raise JPXParseError(f"openpyxl parse failed: {e!r}") from e
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:  # noqa: BLE001
                pass

    if not matrix:
        raise JPXSchemaError("empty sheet (no header row)")

    return parse_rows_from_sheet(_MatrixSheet(matrix))


def parse_jpx_workbook_bytes(content: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """default parse path: bytesシグネチャからxls/xlsxを判定し、対応する
    parserへdispatchする。拡張子ではなくbytesで判定する（§9）。

    未知シグネチャ・HTMLエラーページはJPXParseErrorへ正規化する
    （fetch段の検証をすり抜けたケースへの二重防御）。"""
    try:
        fmt = detect_workbook_format(content)
    except JPXFetchError as e:
        raise JPXParseError(f"workbook signature detection failed: {e}") from e

    if fmt == "xls":
        return parse_jpx_xls_bytes(content)
    return parse_jpx_xlsx_bytes(content)


def detect_duplicate_codes(rows: list[dict[str, Any]]) -> list[str]:
    """code重複を検出する（順序維持、重複した2回目以降のcodeを列挙）。"""
    seen: set[str] = set()
    dupes: list[str] = []
    for row in rows:
        c = row["code"]
        if c in seen:
            dupes.append(c)
        seen.add(c)
    return dupes


def _source_as_of_iso(rows: list[dict[str, Any]]) -> str | None:
    """JPX「日付」列をISO date文字列へ。xlrdはfloat(20260630.0)、openpyxlは
    int(20260630)を返し、稀にdatetime/dateセルの可能性もあるため全て受ける。
    先頭の有効な値を採用する（同一ファイル内で単一のsourceAsOfのため）。"""
    for row in rows:
        raw = row.get("source_as_of_raw")
        if isinstance(raw, bool):
            continue
        if isinstance(raw, datetime):
            return raw.date().isoformat()
        if isinstance(raw, date):
            return raw.isoformat()
        if isinstance(raw, (int, float)):
            s = str(int(raw))
            if len(s) == 8:
                return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"
    return None


# ---------------------------------------------------------------------------
# Eligibility v1
# ---------------------------------------------------------------------------


# JPX/SICC公式の証券コード新設ルールにおける4文字specific-name code
# （普通株式相当。5桁優先株/種類株式=4文字specific-name code+1文字
# reserved codeの構造は別——is_preferred_or_class_share()参照、§12）の
# 位置別文字種契約（P5-B005-R4 Blocker A: R3の`[1-9][0-9A-Z]{3,4}`は
# 桁数のみで英字位置・除外英字を一切見ておらず、1ABC/12A4/123B/999Z等の
# 非正規formatを誤って受理していた）:
#   位置1: 数字のみ（既存JPX code authorityと同じ1-9。0始まりのcodeは
#          実データに現れない、P5-B005-R3 §12を踏襲）
#   位置2: 数字 または 割当許可されたASCII大文字英字
#   位置3: 数字のみ
#   位置4: 数字 または 割当許可されたASCII大文字英字
# JPX/SICCが証券codeへの割当から除外している英字（紛らわしい形状の文字を
# 避けるための公式除外）: B, E, I, O, Q, V, Z。
# 実データ（1301, 7203等の数字のみ、166A, 285A等の位置2/4英字混在）・
# 将来の公式割当例（1A00等、位置2英字）のいずれとも整合する
# （2026-07-14実データのalphanumeric codeは166A/285Aの2件のみだが、
# この2件を正しく受理しつつ非正規formatを拒否するのが本ruleの目的——
# 「今日時点の1552件に含まれる」ことをvalidityの定義にしてはならない、
# P5-B005-R4 §10）。
_JPX_EXCLUDED_LETTERS = frozenset("BEIOQVZ")
_JPX_PERMITTED_DIGIT_OR_LETTER = "".join(
    ch for ch in "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if ch not in _JPX_EXCLUDED_LETTERS
)
_JPX_CODE_SHAPE_RE = re.compile(
    rf"^[1-9][{_JPX_PERMITTED_DIGIT_OR_LETTER}][0-9][{_JPX_PERMITTED_DIGIT_OR_LETTER}]$"
)


def _is_canonical_jpx_code(code: Any) -> bool:
    """codeがlive parse authorityの生成しうる正規形と一致するか。"""
    return isinstance(code, str) and bool(_JPX_CODE_SHAPE_RE.fullmatch(code))


def is_preferred_or_class_share(code: str) -> bool:
    """5桁コードは優先株/種類株式を示す（JPXコード命名規則、
    2026-07-14実データで7件全て「◯◯優先株式」「◯◯種類株式」名称と一致確認済み）。

    契約: 4桁コード（数字のみ・英字混在いずれも）は普通株式として保持する。
    4桁+英字のcode（例: 166A, 285A）は2024年以降の新規上場企業に割り当てられる
    正式な普通株式コードであり、eligibility上除外しない。
    判定は桁数のみで行い、文字種（数字/英字混在）は見ない。
    """
    return len(code) == 5


def apply_eligibility(
    rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, int], list[dict[str, Any]]]:
    """v1 eligible universe filter。

    原則:
      1. 「市場・商品区分」が"プライム（内国株式）"と厳密一致する行のみ採用。
         この1回の厳密一致で以下が同時に達成される
         （2026-07-14実データの区分値10種のうちこれ以外の9種を全て除外）:
           - Prime（Standard/Growth除外）
           - domestic common stocks相当（外国株式区分を除外）
           - ETF/ETN除外・REIT除外・PRO Market除外・出資証券除外
      2. 5桁コード（優先株/種類株式）を除外する（is_preferred_or_class_share）。
         4桁の英字混在codeは普通株式として残す。

    戻り値: (eligible_rows, segment_counts, filters_applied)。
    filters_appliedは各段階のcountを明示するリスト
    （1559 vs 1560差・各filter段階の透明性のため）。
    """
    segment_counts: dict[str, int] = {}
    for row in rows:
        segment_counts[row["market_segment"]] = segment_counts.get(row["market_segment"], 0) + 1

    stage_source = {"stage": FILTER_STAGE_SOURCE_ROWS, "count": len(rows)}

    prime_domestic = [r for r in rows if r["market_segment"] == MARKET_SEGMENT_PRIME_DOMESTIC]
    stage_market = {
        "stage": FILTER_STAGE_MARKET_SEGMENT_PRIME_DOMESTIC,
        "count": len(prime_domestic),
        "excluded_segment_counts": {
            k: v for k, v in segment_counts.items() if k != MARKET_SEGMENT_PRIME_DOMESTIC
        },
    }

    preferred_excluded = [r for r in prime_domestic if is_preferred_or_class_share(r["code"])]
    eligible = [r for r in prime_domestic if not is_preferred_or_class_share(r["code"])]
    stage_preferred = {
        "stage": FILTER_STAGE_EXCLUDE_PREFERRED_OR_CLASS,
        "count": len(eligible),
        "excluded_count": len(preferred_excluded),
        "excluded_codes": sorted(r["code"] for r in preferred_excluded),
    }

    filters_applied = [stage_source, stage_market, stage_preferred]
    return eligible, segment_counts, filters_applied


# ---------------------------------------------------------------------------
# Cache（internal only。public/dataへは一切コピーしない）
# ---------------------------------------------------------------------------


def _parse_iso(raw: Any) -> datetime | None:
    if not isinstance(raw, str) or not raw:
        return None
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except ValueError:
        return None


def _cache_item_valid(item: Any) -> bool:
    """cache items内の1要素が(code, name, sector)契約を満たすか。

    list/tupleでlen==3、code/nameは非空str、sectorはstr（空可）であることを
    要求する。int要素（items=[123]）やstr要素（items=["garbage-item"]）、
    長さ不一致、code/name欠損はいずれも不正として拒否する。"""
    if not isinstance(item, (list, tuple)):
        return False
    if len(item) != 3:
        return False
    code, name, sector = item
    if not isinstance(code, str) or not code:
        return False
    if not isinstance(name, str) or not name:
        return False
    if not isinstance(sector, str):
        return False
    return True


def _cache_payload_valid(payload: Any) -> bool:
    """cacheファイルがschema上有効か（corruption検出）。

    items自体がlistであることに加え、各要素が_cache_item_valid契約を
    満たすことまで深層検証する。malformed itemsを持つcacheはload_cache()で
    Noneとなり、fallback側（_cache_to_result等）が新たな例外源にならない。"""
    if not isinstance(payload, dict):
        return False
    if payload.get("schemaKind") != CACHE_SCHEMA_KIND:
        return False
    items = payload.get("items")
    if not isinstance(items, list):
        return False
    if not all(_cache_item_valid(item) for item in items):
        return False
    if not isinstance(payload.get("row_count"), int):
        return False
    if not isinstance(payload.get("fetched_at"), str) or _parse_iso(payload["fetched_at"]) is None:
        return False
    fmt = payload.get("workbook_format")
    if fmt is not None and fmt not in ("xls", "xlsx"):
        return False
    return True


def load_cache(path: Path = CACHE_PATH) -> dict[str, Any] | None:
    """既存cacheを読み込む。存在しない/corrupt/schema不正ならNone
    （corrupt cacheはfallback候補として使わない）。"""
    if not path.exists():
        return None
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError, UnicodeDecodeError):
        return None
    return raw if _cache_payload_valid(raw) else None


def save_cache(payload: dict[str, Any], path: Path = CACHE_PATH) -> None:
    """cacheを同一ディレクトリの一時ファイルへ書き切ってからatomic replaceする。
    書き込み中断・OSError発生時も、置換前の既存last-good cacheは無傷のまま
    残る（部分書き込みが直接pathへ反映されることはない）。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(path.name + ".tmp")
    try:
        tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp_path.replace(path)
    except BaseException:
        tmp_path.unlink(missing_ok=True)
        raise


def _cache_age_hours(cache_payload: dict[str, Any], now: datetime) -> float:
    fetched = _parse_iso(cache_payload.get("fetched_at"))
    if fetched is None:
        return float("inf")
    if now.tzinfo is None:
        now = now.replace(tzinfo=timezone.utc)
    return (now - fetched).total_seconds() / 3600


def _cache_authority_valid(payload: dict[str, Any], now: datetime) -> bool:
    """load_cache()の構造的validation（corruption検出）を通過したcache
    payloadに対し、last-good authorityとして実際に使ってよいかを再検証する
    （P5-B005-R2: RESTORED_CACHE_REVALIDATED）。

    live provider（get_jpx_universe()のsave_cache()呼び出し箇所）が
    実際に書きうる契約——source/universe_id/eligibility後のitems/
    apply_eligibility()由来のsegment_counts・filters_applied——を正とし、
    構造的にparse可能でもこの契約から外れるpayload（古すぎる/未来日時/
    別source/別universe_id/重複code/floor未満/row_countとの矛盾/
    改竄されたsegment_counts・filters_applied）はいずれもfalseを返す。
    load_cache()自体（schema/型のcorruption検出）は変更しない——単一の
    canonical validatorとしてここに集約し、load_cache()は構造層、
    本関数は意味論層を担う。"""
    # provenance: このprovider自身が書いたcacheであることを要求する。
    if payload.get("source") != SOURCE_IDENTIFIER:
        return False
    if payload.get("universe_id") != UNIVERSE_ID:
        return False

    # freshness: 10年前cache・未来日時cacheをclock-skew tolerance付きで拒否。
    fetched = _parse_iso(payload.get("fetched_at"))
    if fetched is None:
        return False
    now_utc = now if now.tzinfo is not None else now.replace(tzinfo=timezone.utc)
    age_hours = (now_utc - fetched).total_seconds() / 3600
    if age_hours > MAX_CACHE_AGE_HOURS:
        return False
    if age_hours < -CACHE_CLOCK_SKEW_TOLERANCE_HOURS:
        return False

    # universe content: cached itemsは既にeligibility適用後のcanonical
    # universeであるべきなので、live pathと同じ制約を再チェックする——
    # cache metadata（stored row_count/eligible相当件数）を信用せず、
    # 実際のitemsから再計算する。
    items = payload.get("items")
    if not isinstance(items, list) or not items:
        return False
    codes = [item[0] for item in items]
    if not all(_is_canonical_jpx_code(c) for c in codes):
        return False  # live parse authorityが生成しえない形状のcode（例: "BAD!"）
    if len(set(codes)) != len(codes):
        return False  # 重複code（canonical形状のみ受理するため正規化の曖昧さはない）
    if any(is_preferred_or_class_share(c) for c in codes):
        return False  # eligibility上除外されるはずの5桁code混入
    if len(items) < MIN_ELIGIBLE_COUNT:
        return False  # below-floor universe

    # count consistency: row_countはraw行数（items=eligible行数の上位集合）
    # のはずなので、負値・非intは無効。eligible件数を上回らない
    # row_countは矛盾（rowsとの不整合）として拒否する。
    row_count = payload.get("row_count")
    if not isinstance(row_count, int) or isinstance(row_count, bool):
        return False
    if row_count < MIN_RAW_ROW_COUNT:
        return False
    if row_count < len(items):
        return False

    # segment_counts: apply_eligibility()はrowsが非空である限り必ず非空dictを
    # 返す（rows中の各行のmarket_segmentを1件ずつ集計するため）。row_countが
    # MIN_RAW_ROW_COUNT floor以上のlive/cache payloadで空dictが正当に
    # 現れることはないため、空dictは無条件で拒否する（P5-B005-R2は誤って
    # 空dictを許容していた——P5-B005-R3 §14で修正）。keyはJPXが実際に
    # 使う市場区分の承認済み集合（APPROVED_MARKET_SEGMENTS）の範囲内のみ、
    # valueはnon-negative int（bool除外）、合計はrow_countと一致、
    # プライム内国株式区分の値はeligible items件数以上を要求する
    # （apply_eligibility()のstage_source.count==len(rows)==row_countと
    # stage_market.count==segment_counts[PRIME_DOMESTIC]の同値性より）。
    segment_counts = payload.get("segment_counts")
    if not isinstance(segment_counts, dict) or not segment_counts:
        return False
    for key, value in segment_counts.items():
        if not isinstance(key, str) or key not in APPROVED_MARKET_SEGMENTS:
            return False
        if not isinstance(value, int) or isinstance(value, bool) or value < 0:
            return False
    if sum(segment_counts.values()) != row_count:
        return False
    prime_count = segment_counts.get(MARKET_SEGMENT_PRIME_DOMESTIC, 0)
    if prime_count < len(items):
        return False

    # filters_applied: apply_eligibility()が実際に返す形——正確に3段、
    # FILTERS_APPLIED_STAGE_ORDERの名前・順序と厳密一致、countは
    # non-negative int（bool除外）かつ段を追うごとに単調非増加、
    # 初段count==row_count、終段count==eligible items件数、
    # 中段（market_segment）count==segment_counts[PRIME_DOMESTIC]
    # ——を要求する（P5-B005-R3 §16/§17: writer由来の意味論関係を検証）。
    filters_applied = payload.get("filters_applied")
    if not isinstance(filters_applied, list):
        return False
    if len(filters_applied) != len(FILTERS_APPLIED_STAGE_ORDER):
        return False
    prev_count: int | None = None
    for stage, expected_name in zip(filters_applied, FILTERS_APPLIED_STAGE_ORDER):
        if not isinstance(stage, dict):
            return False
        if stage.get("stage") != expected_name:
            return False
        count = stage.get("count")
        if not isinstance(count, int) or isinstance(count, bool) or count < 0:
            return False
        if prev_count is not None and count > prev_count:
            return False  # 段を追うごとに単調非増加でなければならない
        prev_count = count
    if filters_applied[0]["count"] != row_count:
        return False
    if filters_applied[1]["count"] != prime_count:
        return False
    if filters_applied[-1]["count"] != len(items):
        return False

    return True


def cache_authority_valid(payload: dict[str, Any], now: datetime) -> bool:
    """_cache_authority_valid()の公開ラッパー。module外（Full Batch
    workflowのJPX cache save-eligibility gate step等）からcanonical cache
    validatorを再利用するために公開する。ロジックの複製ではなく単一
    canonical実装への薄いshim（P5-B005-R3 §11: single writer/validator
    contract）。"""
    return _cache_authority_valid(payload, now)


# ---------------------------------------------------------------------------
# Current-run cache attestation（P5-B005-R4 §17-22。ephemeral、actions/cache
# には一切persistしない——ATTESTATION_PATH参照）。
# ---------------------------------------------------------------------------


def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _remove_attestation(path: Path = ATTESTATION_PATH) -> None:
    """新しいrunのJPX取得を試みる前に、前run由来のattestationを必ず削除する
    （§20）。live→cache fallback→seed fallbackのいずれの経路を辿っても、
    このrunがちょうど書いたattestationだけが有効という不変条件を保つ。
    削除自体の失敗（存在しない・権限等）はfallback chainを妨げない。"""
    try:
        path.unlink(missing_ok=True)
    except OSError:
        pass


def _write_attestation(
    *,
    run_token: str,
    cache_path: Path,
    source: str,
    fetched_at: str,
    eligible_count: int,
    attestation_path: Path = ATTESTATION_PATH,
) -> None:
    """live JPX取得成功＋canonical cache書き込み直後にのみ呼ばれる。
    cache_sha256はcache_path書き込み"後"に実ファイルbytesを読み直して
    計算する（§18: hashはfinal canonical cache fileが書かれた後に計算
    しなければならない——in-memory payloadのjson.dumps結果を仮定しない）。
    save_cache()と同じ一時ファイル→atomic replaceパターンで書き込む。"""
    cache_sha256 = _sha256_bytes(cache_path.read_bytes())
    payload = {
        "schemaKind": ATTESTATION_SCHEMA_KIND,
        "run_token": run_token,
        "cache_sha256": cache_sha256,
        "source": source,
        "fetched_at": fetched_at,
        "eligible_count": eligible_count,
    }
    attestation_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = attestation_path.with_name(attestation_path.name + ".tmp")
    try:
        tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp_path.replace(attestation_path)
    except BaseException:
        tmp_path.unlink(missing_ok=True)
        raise


def _attestation_payload_valid(payload: Any) -> bool:
    """attestationファイルがschema上有効か（corruption/tamper検出の構造層。
    意味論的なcache/run_tokenとの一致はjpx_cache_save_eligible()が担う）。"""
    if not isinstance(payload, dict):
        return False
    if payload.get("schemaKind") != ATTESTATION_SCHEMA_KIND:
        return False
    run_token = payload.get("run_token")
    if not isinstance(run_token, str) or not run_token:
        return False
    cache_sha256 = payload.get("cache_sha256")
    if not isinstance(cache_sha256, str) or len(cache_sha256) != 64:
        return False
    if any(c not in "0123456789abcdef" for c in cache_sha256):
        return False
    if not isinstance(payload.get("source"), str) or not payload["source"]:
        return False
    if not isinstance(payload.get("fetched_at"), str) or _parse_iso(payload["fetched_at"]) is None:
        return False
    eligible_count = payload.get("eligible_count")
    if not isinstance(eligible_count, int) or isinstance(eligible_count, bool) or eligible_count < 0:
        return False
    return True


def load_attestation(path: Path = ATTESTATION_PATH) -> dict[str, Any] | None:
    """current-run attestationを読み込む。存在しない/corrupt/schema不正なら
    None（load_cache()と同じcorruption-safe契約）。"""
    if not path.exists():
        return None
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError, UnicodeDecodeError):
        return None
    return raw if _attestation_payload_valid(raw) else None


def jpx_cache_save_eligible(
    *,
    candidates_meta: dict[str, Any],
    expected_run_token: str,
    cache_payload: dict[str, Any] | None,
    cache_bytes: bytes | None,
    attestation: dict[str, Any] | None,
    now: datetime,
) -> tuple[bool, str]:
    """P5-B005-R4 §22: Full Batch workflowのJPX cache save-eligibility gate
    が使う単一canonical判定（cache_authority_validと同じ「複製ではなく
    module外から再利用する単一実装」契約、§11踏襲）。

    aggregate _meta.pipelinePathは使わない（§16: prescreenだけがcache
    fallbackへ回ってもJPX自体はlive成功していうるケース＝§27の
    JPX_LIVE_AND_PRESCREEN_FALLBACK_STILL_ELIGIBLEを誤ってfalseにする
    ため）。代わりにJPX固有のprovenance
    （_meta.universeProvenance.jpxFallbackUsed/jpxSource/jpxEligibleCount、
    既存whole_market_universe_provider()が同一run内で書く値）＋canonical
    cache authority＋current-run attestationのcontent binding
    （run_token一致・SHA256一致・source/eligible_count/fetched_at一致）の
    全てを要求する。いずれか1つでも欠けたり食い違えばFalseを返す
    （fail closed）。戻り値の理由文字列はworkflow step診断ログとテスト
    双方から使う。"""
    if not isinstance(expected_run_token, str) or not expected_run_token:
        return False, "expected_run_token is empty"
    if candidates_meta.get("runToken") != expected_run_token:
        return False, "candidates_stocks.json _meta.runToken does not match this run"

    provenance = candidates_meta.get("universeProvenance")
    if not isinstance(provenance, dict):
        return False, "candidates_stocks.json _meta.universeProvenance is missing"
    if provenance.get("jpxFallbackUsed") is not False:
        return False, (
            "universeProvenance.jpxFallbackUsed is not a fresh live JPX acquisition"
        )
    if provenance.get("jpxSource") != SOURCE_IDENTIFIER:
        return False, "universeProvenance.jpxSource is not the expected JPX source identifier"

    if cache_payload is None:
        return False, "cache file does not exist or is not valid JSON"
    if not cache_authority_valid(cache_payload, now):
        return False, "cache payload fails canonical authority validation"

    if provenance.get("jpxEligibleCount") != len(cache_payload.get("items", [])):
        return False, "universeProvenance.jpxEligibleCount does not match cache items"

    if attestation is None:
        return False, "current-run cache attestation is missing or invalid"
    if attestation.get("run_token") != expected_run_token:
        return False, "attestation run_token does not match this run"
    if cache_bytes is None or attestation.get("cache_sha256") != _sha256_bytes(cache_bytes):
        return False, "attestation cache_sha256 does not match the cache file bytes"
    if attestation.get("source") != cache_payload.get("source"):
        return False, "attestation source does not match cache source"
    if attestation.get("eligible_count") != len(cache_payload.get("items", [])):
        return False, "attestation eligible_count does not match cache items"
    if attestation.get("fetched_at") != cache_payload.get("fetched_at"):
        return False, "attestation fetched_at does not match cache fetched_at"

    return True, "eligible"


def _cache_to_result(cache_payload: dict[str, Any], now: datetime) -> JPXUniverseResult:
    """last-good cacheをJPXUniverseResultへ変換する。fallbackUsed=True、
    timestampはcache生成時のfetched_atをそのまま使う（偽装しない）。"""
    items = [tuple(x) for x in cache_payload["items"]]
    return JPXUniverseResult(
        universe_id=cache_payload.get("universe_id", UNIVERSE_ID),
        items=items,
        source=cache_payload.get("source", SOURCE_IDENTIFIER),
        source_identifier=SOURCE_IDENTIFIER,
        fetched_at=cache_payload["fetched_at"],
        source_as_of=cache_payload.get("source_as_of"),
        row_count=cache_payload["row_count"],
        eligible_count=len(items),
        segment_counts=cache_payload.get("segment_counts", {}),
        filters_applied=cache_payload.get("filters_applied", []),
        fallback_used=True,
        cache_age_hours=_cache_age_hours(cache_payload, now),
        dropped_rows=[],
        workbook_format=cache_payload.get("workbook_format"),
    )


def seed_list_v1_fallback(now: datetime) -> JPXUniverseResult:
    """live fetch異常・valid cacheなしの最終fallback。既存
    data/build_candidates_stocks.pyのSEED_LIST(41銘柄)をそのまま使う
    （既存41 seed production挙動と同一の縮退universe）。"""
    from data.build_candidates_stocks import SEED_LIST

    now_iso = now.isoformat()
    return JPXUniverseResult(
        universe_id=FALLBACK_UNIVERSE_ID,
        items=list(SEED_LIST),
        source="data/build_candidates_stocks.py::SEED_LIST",
        source_identifier=FALLBACK_UNIVERSE_ID,
        fetched_at=now_iso,
        source_as_of=None,
        row_count=len(SEED_LIST),
        eligible_count=len(SEED_LIST),
        segment_counts={},
        filters_applied=[],
        fallback_used=True,
        cache_age_hours=None,
        dropped_rows=[],
        workbook_format=None,
    )


# ---------------------------------------------------------------------------
# Main entrypoint: fetch → validate → eligibility → cache/fallback chain
# ---------------------------------------------------------------------------


def get_jpx_universe(
    now: datetime | None = None,
    fetch_fn: Any = fetch_jpx_workbook_bytes,
    parse_fn: Any = parse_jpx_workbook_bytes,
    cache_path: Path = CACHE_PATH,
    run_token: str | None = None,
    attestation_path: Path = ATTESTATION_PATH,
) -> JPXUniverseResult:
    """JPX universe providerの主エントリポイント。failure chain:
      1. live fetch成功（listing page discover → workbook download →
         signature検証）→ validate（schema/duplicate/row-count/
         eligible-count guard）→ eligible universe（fallbackUsed=False）
      2. fetch/parse/schema/duplicate/row-count/eligible-count異常
         → valid last-good cache（fallbackUsed=True, cacheAgeHours>=0）
      3. valid cacheなし → seed_list_v1 fallback
         （fallbackUsed=True, cacheAgeHours=None）

    corrupt cacheは使わない。stale状態はcacheAgeHoursで可視化する
    （fallback時にtimestampを偽装しない）。guard失敗時はsave_cache()に
    到達しないため、last-good cacheが異常結果で上書きされることはない
    （first-runでtruncated sourceをbaseline化することも防ぐ）。

    integrity guardは4段階（raw row絶対floor→raw row前回比ratio→
    eligible絶対floor→eligible前回比ratio）で、raw rowは正常でも
    market label drift等でeligibleだけ崩壊するケース（F1）と、
    前回cacheが無いfirst-runでtruncated sourceがbaseline化される
    ケース（F7）の両方を捕捉する。

    fetch_fn/parse_fnはテスト用のdependency injectionポイント
    （fetch_fn: () -> bytes、parse_fn: (bytes) -> (rows, dropped)）。
    defaultはlisting-page discovery経路（fetch_jpx_workbook_bytes）と
    signature-dispatch parse（parse_jpx_workbook_bytes, xls/xlsx両対応）。

    run_tokenはFull Batchの既存run-token（build_candidates_stocks.py
    main()の--run-tokenをそのまま透過）。live JPX取得成功＋canonical
    cache書き込みが完了した場合のみ、current-run attestation
    （attestation_path、既定ATTESTATION_PATH）をこのtokenで書く
    （P5-B005-R4 §17-22）。Noneの場合はattestationを一切書かない
    （offline test・ad-hoc実行等、run-tokenが存在しない呼び出し）。
    どの経路（live成功/cache fallback/seed fallback）でも、新しい取得を
    試みる前に必ず前run由来のstale attestationを削除する（§20）。
    """
    if now is None:
        now = datetime.now(timezone.utc)

    _remove_attestation(attestation_path)

    cache = load_cache(cache_path)
    if cache is not None and not _cache_authority_valid(cache, now):
        # 構造的にparse可能でもlast-good authorityの契約（provenance/
        # freshness/universe content/count consistency）を満たさない
        # cacheは、guard比較にも_cache_to_result()のfallback候補にも使わない
        # （P5-B005-R2: RESTORED_CACHE_REVALIDATED）。
        print(
            "[jpx_universe_provider] cached last-good universe failed authority "
            "revalidation, discarding",
            file=sys.stderr,
        )
        cache = None
        # defense in depth（P5-B005-R3 §27）: authority-invalidと判定された
        # restored cache fileはbest-effortで削除する。これは
        # Full BatchのCACHE_SAVE_CONDITION gate（現在-run live provenance
        # 相関）を置き換えるものではない——このprocessが仮にcrash/skipして
        # ファイルが残っても、workflow側のsave-eligibility gateが単独で
        # 安全である設計を維持する。削除自体の失敗（権限・並行削除等）は
        # fallback chainを妨げない。
        try:
            cache_path.unlink(missing_ok=True)
        except OSError:
            pass

    try:
        content = fetch_fn()
        rows, dropped = parse_fn(content)

        try:
            workbook_format: str | None = detect_workbook_format(content)
        except (JPXFetchError, TypeError):
            # injected fake fetch（b"irrelevant"等）では判定不能。provenanceは
            # Noneのまま——parse自体は成功しているので結果は有効。
            workbook_format = None

        dupes = detect_duplicate_codes(rows)
        if dupes:
            raise JPXSchemaError(f"duplicate codes detected: {dupes[:10]}")

        row_count = len(rows)
        if row_count < MIN_RAW_ROW_COUNT:
            raise JPXRowCountGuardError(
                f"row_count {row_count} is below absolute floor MIN_RAW_ROW_COUNT="
                f"{MIN_RAW_ROW_COUNT}"
            )
        if cache is not None:
            prev_row_count = cache.get("row_count", 0)
            if prev_row_count > 0 and row_count < prev_row_count * ROW_COUNT_MIN_RATIO:
                raise JPXRowCountGuardError(
                    f"row_count {row_count} < {ROW_COUNT_MIN_RATIO * 100:.0f}% "
                    f"of previous good row_count {prev_row_count}"
                )

        eligible, segment_counts, filters_applied = apply_eligibility(rows)

        eligible_count = len(eligible)
        if eligible_count < MIN_ELIGIBLE_COUNT:
            raise JPXEligibleCountGuardError(
                f"eligible_count {eligible_count} is below absolute floor "
                f"MIN_ELIGIBLE_COUNT={MIN_ELIGIBLE_COUNT}"
            )
        if cache is not None:
            prev_eligible_count = len(cache.get("items", []))
            if prev_eligible_count > 0 and eligible_count < prev_eligible_count * ELIGIBLE_COUNT_MIN_RATIO:
                raise JPXEligibleCountGuardError(
                    f"eligible_count {eligible_count} < {ELIGIBLE_COUNT_MIN_RATIO * 100:.0f}% "
                    f"of previous good eligible_count {prev_eligible_count}"
                )

        source_as_of = _source_as_of_iso(rows)
        items = [(r["code"], r["name"], r["sector"]) for r in eligible]

        result = JPXUniverseResult(
            universe_id=UNIVERSE_ID,
            items=items,
            source=SOURCE_IDENTIFIER,
            source_identifier=SOURCE_IDENTIFIER,
            fetched_at=now.isoformat(),
            source_as_of=source_as_of,
            row_count=row_count,
            eligible_count=len(items),
            segment_counts=segment_counts,
            filters_applied=filters_applied,
            fallback_used=False,
            cache_age_hours=0.0,
            dropped_rows=dropped,
            workbook_format=workbook_format,
        )

        save_cache({
            "schemaKind": CACHE_SCHEMA_KIND,
            "universe_id": result.universe_id,
            "items": [list(x) for x in items],
            "source": result.source,
            "fetched_at": result.fetched_at,
            "source_as_of": result.source_as_of,
            "row_count": result.row_count,
            "segment_counts": result.segment_counts,
            "filters_applied": result.filters_applied,
            "workbook_format": result.workbook_format,
        }, cache_path)

        if run_token is not None:
            # §18: hashはcanonical cache fileが書かれた"後"に実bytesを
            # 読み直して計算する（_write_attestation内部）。
            _write_attestation(
                run_token=run_token,
                cache_path=cache_path,
                source=result.source,
                fetched_at=result.fetched_at,
                eligible_count=result.eligible_count,
                attestation_path=attestation_path,
            )

        return result

    except (
        JPXFetchError,
        JPXParseError,
        JPXSchemaError,
        JPXRowCountGuardError,
        JPXEligibleCountGuardError,
    ) as e:
        print(f"[jpx_universe_provider] live fetch/validate failed: {e!r}", file=sys.stderr)
        if cache is not None:
            print("[jpx_universe_provider] falling back to last-good cache", file=sys.stderr)
            return _cache_to_result(cache, now)
        print("[jpx_universe_provider] no valid cache, falling back to seed_list_v1", file=sys.stderr)
        return seed_list_v1_fallback(now)
